
import openpyxl
import textwrap
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference
import yfinance as yf
from datetime import datetime

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1B2A4A"
MID_BLUE    = "2F5496"
LIGHT_BLUE  = "BDD7EE"
YELLOW_IN   = "FFF2CC"   # input cells
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"
GREEN_BG    = "E2EFDA"
RED_BG      = "FFC7CE"
GREEN_FONT  = "375623"
RED_FONT    = "9C0006"

def side(style="thin"):
    return Side(border_style=style, color="AAAAAA")

THIN_BORDER = Border(left=side(), right=side(), top=side(), bottom=side())

def hdr_border():
    return Border(left=side(), right=side(), top=side("medium"), bottom=side("medium"))

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Category definitions ──────────────────────────────────────────────────────
CATEGORIES = [
    {
        "name": "📊 Dashboard",
        "tab_color": "1B2A4A",
        "color": "1B2A4A",
        "description": "",
        "stocks": [],
    },
    {
        "name": "⚛️ Quantum Computing",
        "tab_color": "7030A0",
        "color": "7030A0",
        "description": "High-risk quantum plays — potential 10× returns within 2 years",
        "stocks": [
            {"ticker": "IONQ",  "company": "IonQ Inc",          "focus": "Quantum computing hardware & software",   "risk": "Very High", "notes": "Next quantum run candidate. The belief: 'Buy IONQ, RGTI, and QBTS — easy 10x within 2 years.' IonQ is the most commercially advanced pure-play quantum computing company, using trapped-ion technology which is considered one of the most promising paths to fault-tolerant quantum computers. They already have paying customers (government, pharma, finance) and are generating real revenue. When the next quantum hype cycle hits, IonQ will be the headline name that institutions pile into. Very speculative — treat as a high-risk asymmetric bet and size position accordingly."},
            {"ticker": "RGTI",  "company": "Rigetti Computing",  "focus": "Quantum computing chips & systems",        "risk": "Very High", "notes": "Next quantum run candidate alongside IONQ and QBTS. Rigetti builds its own quantum processors and offers cloud-based quantum computing access. It's a smaller, scrappier name than IonQ, which means more volatility but also more upside if they execute. The company has been building out its QPU (quantum processing unit) roadmap and targeting commercial applications in chemistry, logistics, and finance. When the next quantum narrative cycle accelerates, RGTI tends to move fast and hard. Pure asymmetric bet — small position, high conviction on timing."},
            {"ticker": "QBTS",  "company": "D-Wave Quantum",     "focus": "Quantum annealing systems",               "risk": "Very High", "notes": "Next quantum run candidate alongside IONQ and RGTI. D-Wave is unique in that it uses quantum annealing rather than gate-based quantum computing, which makes it commercially viable TODAY for specific optimization problems — logistics, scheduling, materials discovery — that classical computers struggle with. It already has real enterprise customers generating recurring revenue. The commercial viability argument is stronger than most quantum names. Easy 10x potential if the next quantum hype cycle materializes. Very speculative — size position accordingly."},
        ],
    },
    {
        "name": "🖥️ Semiconductors & Compute",
        "tab_color": "0070C0",
        "color": "0070C0",
        "description": "Core semiconductor & compute names — stability + steady compounders",
        "stocks": [
            {"ticker": "NVDA",  "company": "NVIDIA Corp",              "focus": "GPUs, AI compute",              "risk": "Medium",      "notes": "The AI compute king and the core stability position in this portfolio. NVIDIA doesn't just make GPUs — it has built the dominant software ecosystem (CUDA) that makes its hardware nearly impossible to replace even if competitors close the hardware gap. Every major hyperscaler (Google, Amazon, Microsoft, Meta) depends on NVIDIA for AI training. As AI spending continues to accelerate, NVIDIA captures the most direct revenue from every dollar spent on AI compute. This is a steady compounder that also has massive upside as the AI buildout continues. Hold as a core long-term position."},
            {"ticker": "AVGO",  "company": "Broadcom Inc",             "focus": "Semiconductors, networking",    "risk": "Medium",      "notes": "Stability and steady compounder — one of the most reliable semiconductor businesses in the world. Broadcom makes the networking chips that connect all the GPUs inside hyperscaler data centers. As AI clusters scale to hundreds of thousands of GPUs, the networking infrastructure connecting them becomes more valuable and more complex. Broadcom also has a massive software revenue stream from its VMware acquisition that provides recurring cash flow stability. Management is world-class at capital allocation. This is a core hold that compounds quietly while the AI buildout accelerates."},
            {"ticker": "LRCX",  "company": "Lam Research",             "focus": "Semiconductor equipment, DRAM", "risk": "Medium",      "notes": "Stability and steady compounder — DRAM equipment leader. Lam Research makes the etch and deposition equipment that builds memory chips (DRAM and NAND). AI requires massive amounts of high-bandwidth memory (HBM), and every HBM chip that gets manufactured requires Lam's tools. This gives Lam direct and durable exposure to the AI memory build-out without the volatility of owning the memory chipmakers directly. The business is highly recurring (installed base + consumables) and grows every time wafer starts increase. A core picks-and-shovels play for the AI infrastructure trade."},
            {"ticker": "AMD",   "company": "Advanced Micro Devices",   "focus": "CPUs, GPUs, AI chips",          "risk": "Medium",      "notes": "CPU maker deep in the AI infrastructure stack and the only credible challenger to NVIDIA in AI accelerators. AMD's MI300X GPU is winning meaningful enterprise AI workloads, particularly inference, where customers want a cheaper alternative to NVIDIA's H100/H200. On the CPU side, AMD's EPYC chips power a huge share of hyperscaler server infrastructure. The AI narrative is also shifting down-stack into CPU makers — AMD benefits from both trends simultaneously. Risk: NVIDIA's software moat (CUDA) is still wide. But as a compounder and a hedge on NVIDIA dominance, AMD belongs in the portfolio."},
            {"ticker": "INTC",  "company": "Intel Corp",               "focus": "CPUs, AI accelerators",         "risk": "Med-High",    "notes": "CPU maker in the AI infrastructure stack and the most controversial turnaround story in semiconductors. Intel lost its manufacturing lead to TSMC and its AI accelerator business to NVIDIA and AMD. The thesis here is a turnaround: new management, massive government subsidies via the CHIPS Act, and a path to reclaiming leading-edge manufacturing through its Intel Foundry Services (IFS) division. If the turnaround works, INTC is deeply undervalued. If it doesn't, the downside is meaningful. This is a higher-risk position — size it as a speculative turnaround bet, not a core hold. The AI infrastructure narrative (CPU makers shifting down-stack) gives it a tailwind even without a full recovery."},
            {"ticker": "ARM",   "company": "Arm Holdings",             "focus": "Chip architecture licensing",   "risk": "Medium",      "notes": "The architecture behind nearly every AI chip that isn't an x86 CPU. ARM licenses its instruction set architecture to Apple, Qualcomm, Amazon (Graviton), Google, NVIDIA, and dozens of others. Every time an AI edge chip, smartphone chip, or custom datacenter chip gets designed and sold, ARM earns a royalty. This is a pure royalty business model — no manufacturing risk, no inventory risk — just IP licensing that scales with every chip shipped globally. As AI moves to the edge (phones, cars, robots), ARM's exposure compounds. The royalty rate per chip is also increasing as designs get more complex. A core compounder."},
        ],
    },
    {
        "name": "⚙️ AI Picks & Shovels",
        "tab_color": "C55A11",
        "color": "C55A11",
        "description": "Smaller-cap NVDA-adjacent setups — high risk, asymmetric upside",
        "stocks": [
            {"ticker": "NVTS",  "company": "Navitas Semiconductor",  "focus": "GaN power chips for AI data centers",             "risk": "High",      "notes": "The NVDA-parallel picks-and-shovels bet for AI power. Navitas makes GaN (gallium nitride) power chips that are radically more efficient than traditional silicon-based power delivery. AI data centers are hitting a hard wall on power density — the amount of electricity you can push into a rack is physically limited by current power conversion technology. GaN solves this by converting power more efficiently and at higher frequencies, enabling denser, cooler, more power-efficient data centers. As hyperscalers build out hundreds of gigawatts of AI compute, NVTS stands to be a critical infrastructure supplier. High-risk, asymmetric upside — the kind of bet that looks obvious in hindsight."},
            {"ticker": "AEHR",  "company": "AEHR Test Systems",       "focus": "Semiconductor burn-in & test equipment",          "risk": "High",      "notes": "Classic shovels-in-a-gold-rush setup. AEHR makes specialized semiconductor burn-in and test equipment used to stress-test chips before they ship. As AI chips become more complex and expensive (each H100 costs $30k+), the cost of shipping a defective chip becomes enormous — which means chip manufacturers are investing heavily in advanced testing. More AI chips manufactured = more testing bottleneck demand = more AEHR revenue. The business is lumpy (large equipment orders), which creates volatility, but the long-term tailwind is durable. A pure picks-and-shovels beneficiary of the AI chip production ramp."},
            {"ticker": "CRDO",  "company": "Credo Technology",        "focus": "High-speed connectivity / SerDes for AI DCs",     "risk": "High",      "notes": "One of the cleanest AI infrastructure picks-and-shovels names available. GPUs are useless if the data can't move between them fast enough. Credo makes high-speed SerDes (serializer/deserializer) connectivity chips that are the plumbing inside AI data centers — connecting GPUs to memory, switches, and storage at speeds that match the compute throughput. As AI clusters scale from thousands to hundreds of thousands of GPUs, the interconnect bandwidth requirement explodes. Credo has direct customer relationships with hyperscalers and is already generating real revenue from the AI buildout. Strong execution, growing bookings, and a durable structural tailwind."},
            {"ticker": "ALGM",  "company": "Allegro MicroSystems",    "focus": "Power & sensing semiconductors",                  "risk": "Med-High",  "notes": "A quiet compounder that is early in its AI rerating narrative. Allegro makes power management and sensing semiconductors used in automotive, industrial, and increasingly AI hardware applications. As AI hardware demands more precise power delivery and thermal sensing, Allegro's chips become more embedded in the data center infrastructure. The automotive exposure (EV power management) provides a separate growth vector. Less hype than the pure-play AI names, which means it hasn't been fully rerated yet — creating a potential catch-up opportunity. More of a steady grower than a moonshot, but with meaningful upside as the narrative broadens."},
            {"ticker": "HIMX",  "company": "Himax Technologies",      "focus": "Display drivers + AR/AI vision chips",            "risk": "High",      "notes": "An often-overlooked ADR with cyclical breakout behavior. Himax makes display driver ICs and AI vision chips used in AR/VR devices, automotive displays, and AI edge sensing applications. When AI edge device cycles accelerate — smartphones, AR glasses, smart cameras — Himax tends to move fast. It's historically undervalued relative to its revenue and has a pattern of sharp re-ratings on earnings beats during upcycles. The AR/AI vision chip exposure is genuinely interesting as AI moves from the cloud to the edge. Higher volatility, but the kind of name that can double in a single cycle. Treat as a tactical position sized for the volatility."},
            {"ticker": "MTSI",  "company": "MACOM Technology",        "focus": "RF, microwave, high-speed analog chips",          "risk": "Med-High",  "notes": "Less hype, more industrial AI backbone. MACOM makes RF, microwave, and high-speed analog chips used in data center interconnects, telecom backbones, and defense systems. As AI data centers demand ever-higher bandwidth between racks and buildings, MACOM's optical and RF interconnect chips become critical infrastructure. The telecom angle also gives it exposure to 5G buildout and the long-term bandwidth expansion required to support AI inference at the edge. Not a headline name, but a steady grower with durable end-market demand. The kind of stock that compounds quietly while the hype goes elsewhere, then re-rates when the market catches up."},
            {"ticker": "AEVA",  "company": "Aeva Technologies",       "focus": "Lidar + sensing for autonomous systems",          "risk": "Very High", "notes": "Pure asymmetric bet — low probability, high upside if they execute. Aeva makes FMCW (frequency-modulated continuous-wave) lidar sensors for autonomous vehicles and robotics. Their approach is technically differentiated — FMCW lidar can detect velocity directly (not just position) and works better in adverse conditions than traditional lidar. The autonomous vehicle and robotics market is massive if it materializes, and Aeva would be a key sensing supplier. Still in early commercialization — revenue is minimal and the path to profitability is long. This is a speculative moonshot, not a core position. Size it as a lottery ticket: accept that it could go to zero, but the upside if robotics/AV take off is 10-20x."},
        ],
    },
    {
        "name": "💡 Photonics & Optics",
        "tab_color": "00B0F0",
        "color": "00B0F0",
        "description": "Optical networking — key bottleneck in AI data center interconnect",
        "stocks": [
            {"ticker": "LITE",  "company": "Lumentum Holdings",  "focus": "Photonics, optical components for AI DCs",  "risk": "High",  "notes": "King of Photonics — and this could be a genuine moonshot. Lumentum makes the optical components (lasers, transceivers, photonic integrated circuits) that transmit data at the speed of light inside and between AI data centers. As AI clusters scale to hundreds of thousands of GPUs spread across multiple buildings, the optical interconnect becomes the critical bottleneck — you can't run copper cables at those distances and speeds. Lumentum is one of the few companies with the technology and scale to supply this infrastructure. The AI tailwind here is massive and underappreciated. Photonics is the next big narrative after GPUs and power — when the market wakes up to the data movement bottleneck, LITE is the headline name. High conviction, high risk, moonshot potential."},
        ],
    },
    {
        "name": "🏗️ DC Builders & Cooling",
        "tab_color": "375623",
        "color": "375623",
        "description": "Builders, materials, and cooling companies powering AI data center construction",
        "stocks": [
            {"ticker": "CAT",   "company": "Caterpillar Inc",       "focus": "Heavy equipment, construction",            "risk": "Low-Med",   "notes": "The actual builder and materials company in the AI data center trade. AI data centers are physical buildings requiring massive earthmoving, foundation work, and construction equipment — and Caterpillar is the dominant supplier of that heavy equipment globally. Every hyperscaler campus that Meta, Google, Microsoft, and Amazon builds requires CAT equipment on site. This is a steadier, lower-volatility way to play the AI infrastructure buildout than the semiconductor names. CAT also benefits from infrastructure spending more broadly (roads, bridges, energy projects) which provides diversification. A solid core position — steady compounder with the AI tailwind as a bonus."},
            {"ticker": "FIX",   "company": "Comfort Systems USA",   "focus": "HVAC, mechanical, electrical for DCs",    "risk": "Medium",    "notes": "Builder AND cooling company — one of the best-positioned picks-and-shovels plays in the AI buildout. Comfort Systems provides HVAC, mechanical, plumbing, and electrical systems for commercial buildings, including AI data centers. Cooling is a MAJOR and growing bottleneck in AI data centers — every GPU cluster generates enormous heat that must be managed. FIX designs and installs the entire MEP (mechanical, electrical, plumbing) system that keeps these facilities running. They have direct contracts with hyperscalers and are seeing demand accelerate as AI data center construction ramps. A quality business with real earnings growth, not a speculative bet."},
            {"ticker": "POWL",  "company": "Powell Industries",     "focus": "Electrical infrastructure for DCs",       "risk": "Medium",    "notes": "Electrical infrastructure builder for AI data centers — the company that builds the switchgear, motor control centers, and power distribution equipment that routes electricity inside data centers and industrial facilities. As AI data centers require ever-larger amounts of power (some campus projects exceed 1 GW), the electrical infrastructure to safely distribute that power becomes a critical bottleneck. Powell has a backlog-heavy business model with long-cycle orders, which means revenue visibility is strong. Less volatile than the semiconductor names because of the order backlog and industrial nature of the business. A solid mid-cap infrastructure pick with a genuine AI tailwind."},
            {"ticker": "VRT",   "company": "Vertiv Holdings",       "focus": "Cooling systems for AI data centers",     "risk": "Medium",    "notes": "The premier AI data center cooling play and one of the most direct beneficiaries of the AI buildout. Vertiv makes thermal management systems (liquid cooling, air cooling, heat exchangers) along with power distribution infrastructure for data centers. Cooling is not optional — you cannot run a GPU cluster without solving the thermal problem, and as GPU power consumption has exploded from 300W to 700W per chip, existing cooling infrastructure is completely inadequate. Every new AI data center that gets built needs Vertiv's systems. The company has multi-year backlog, growing margins, and is adding capacity to meet demand. A high-conviction hold in the AI infrastructure basket."},
        ],
    },
    {
        "name": "⚡ Power & Energy",
        "tab_color": "ED7D31",
        "color": "ED7D31",
        "description": "Power generation, grid maintenance & nuclear energy for the AI buildout",
        "stocks": [
            {"ticker": "GEV",   "company": "GE Vernova",      "focus": "Power generation & grid equipment",        "risk": "Medium",    "notes": "Stability and steady compounder — the premier power generation and grid equipment company spun out of GE. GE Vernova makes gas turbines, wind turbines, grid software, and electrification equipment. AI data centers are consuming power at a rate that is straining the US electrical grid, and new power generation capacity takes years to build. GE Vernova is one of the few companies with the scale and installed base to actually supply that new capacity fast. Their gas turbine backlog is multi-year and growing. This is not a speculative bet — it's a high-quality industrial compounder with a very real and durable demand tailwind from AI power needs. A core long-term hold."},
            {"ticker": "PWR",   "company": "Quanta Services", "focus": "Power grid infrastructure & maintenance",  "risk": "Medium",    "notes": "Power grid infrastructure and maintenance — a growing bottleneck for the AI buildout. Quanta Services is the largest specialty contractor for electric power infrastructure in North America. They build and maintain power lines, substations, and grid connections that are essential to delivering electricity to AI data centers. Every new data center campus needs a grid connection, often requiring new transmission lines and substations — all of which Quanta builds. The company has multi-year visibility through its backlog and framework agreements with utilities. A steady, high-quality compounder that benefits directly and durably from the AI power demand surge. Lower volatility than the semiconductor names with a comparably strong tailwind."},
            {"ticker": "BE",    "company": "Bloom Energy",    "focus": "Fuel cell power generation",              "risk": "High",      "notes": "Moonshot potential if hydrogen and fuel cell technology takes off — and the AI power crisis is accelerating the timeline. Bloom Energy makes solid oxide fuel cells that generate electricity on-site without connecting to the grid. For AI data centers that need reliable, clean, on-site power (especially in areas with grid congestion), Bloom's solution is genuinely compelling. They're already selling to data centers and have partnerships with major tech companies. The bull case: AI power demand is so acute that on-site generation becomes mandatory, and Bloom is uniquely positioned. The bear case: fuel cells are expensive and hydrogen infrastructure is still nascent. Worth holding a position — the upside if they win even a small share of the AI power market is very large."},
            {"ticker": "OKLO",  "company": "Oklo Inc",        "focus": "Nuclear SMRs + nuclear fuel waste",        "risk": "Very High", "notes": "Nuclear SMRs and nuclear fuel waste company with massive moonshot potential — but big regulatory and execution risk. Oklo is building small modular nuclear reactors (SMRs) that could provide clean, always-on baseload power for AI data centers at scale. Several major tech companies (including Microsoft) have already signed nuclear power agreements, signaling that nuclear is being taken seriously as an AI power solution. Oklo also has a fuel recycling angle (using spent nuclear fuel) that could be a significant business in itself. The regulatory path is long and uncertain, and Oklo has not yet built a commercial reactor. This is a small-% portfolio position — accept that it could go to zero, but the upside if nuclear SMRs get commercially deployed is enormous and the valuation reflects very little of that potential today."},
        ],
    },
    {
        "name": "☁️ Neoclouds & AI Software",
        "tab_color": "7030A0",
        "color": "5A2E8C",
        "description": "Next-gen cloud infrastructure and AI software platforms",
        "stocks": [
            {"ticker": "NBIS",  "company": "Nebius Group",           "focus": "Neocloud + Clickhouse + AV Ride",       "risk": "High",      "notes": "The leading Neocloud and one of the most interesting companies in this portfolio. Nebius is a modern Amazon/Google-style company: it has a primary product (GPU cloud computing for AI companies that can't get capacity from AWS/Azure/GCP) AND multiple other revenue streams that could each become massive independently — Clickhouse (high-performance analytics database used by thousands of companies), AV Ride (autonomous vehicle tech), and others. The neocloud thesis: hyperscalers are capacity-constrained and have long waitlists for GPU compute, creating a massive opening for alternative cloud providers. NBIS is the best-positioned name to fill that gap. I believe this company could become a future Hyperscaler in the same way Amazon AWS or Google Cloud became defining infrastructure businesses. High conviction, long-term hold — this is the kind of company you look back on in 10 years and wish you owned more of."},
            {"ticker": "PLTR",  "company": "Palantir Technologies",  "focus": "AI software, government AI stack",      "risk": "Med-High",  "notes": "So integral to the entire US government AI stack that it is effectively a national security asset. Palantir's software (Gotham for government, Foundry for enterprise, AIP for AI) is deeply embedded in the Department of Defense, intelligence agencies, and dozens of federal programs. You don't rip out Palantir — agencies build entire workflows on top of it. The AI Platform (AIP) is now being adopted by commercial enterprises as well, opening a second growth vector. The $1.5T defense budget and accelerating government AI spending is a direct tailwind. Yes, the valuation is demanding, but for a company this strategically embedded, the premium is arguably justified. An outlier name outside the data center trade that belongs in a long-term AI portfolio."},
        ],
    },
    {
        "name": "🛡️ Defense & Drones",
        "tab_color": "595959",
        "color": "595959",
        "description": "Defense technology and drone companies — $1.5T defense budget tailwind",
        "stocks": [
            {"ticker": "ONDS",  "company": "Ondas Holdings",  "focus": "Drones and defense tech",  "risk": "Very High", "notes": "Drones and defense tech — very speculative but with a massive catalyst behind it. The US government just announced a $1.5 trillion defense budget with $56 billion going specifically to drones and counter-drones through the DAWG (Drone and Autonomous Warfare Group) program. Ondas is one of the few small-cap companies with actual drone technology, recent acquisitions in the space, and partnerships that position them to compete for these contracts. Their March 2026 acquisitions and partnerships were specifically designed to make them a stronger DAWG program competitor. This is not a guaranteed win — it's a very speculative bet on a small company winning government contracts in a highly competitive space. But if they win even a small slice of the $56B drone budget, the upside from current levels is enormous. Size it appropriately — this could go to zero or it could 5-10x."},
        ],
    },
    {
        "name": "🚀 Space Economy",
        "tab_color": "002060",
        "color": "002060",
        "description": "Space infrastructure — long-duration compounder with SpaceX IPO catalyst",
        "stocks": [
            {"ticker": "RKLB",  "company": "Rocket Lab USA",  "focus": "Space economy & launch infrastructure",  "risk": "High", "notes": "Space economy and space infrastructure — huge future TAM over the next decade and beyond. Rocket Lab is the second most active launch company in the world (after SpaceX) and is the only credible alternative for small and medium satellite launches. Their Electron rocket has an exceptional launch record, and the larger Neutron rocket (in development) is designed to compete with SpaceX's Falcon 9 for medium payloads. The CEO (Peter Beck) is brilliant and has consistently delivered on technical milestones. Near-term catalyst: the SpaceX IPO (expected June 2026) will drive massive media attention to the space economy, and RKLB will benefit as the most investable public proxy. Longer term, Rocket Lab is building a vertically integrated space company (satellites, components, software, launch) which gives them multiple revenue streams. Speculative but with massive compounding upside if they execute over the next 5-10 years."},
        ],
    },
    {
        "name": "🔭 Long Term Watchlist",
        "tab_color": "2E75B6",
        "color": "2E75B6",
        "description": "Long-term investment research candidates under active evaluation",
        "stocks": [
            {"ticker": "RBRK",  "company": "Rubrik Inc",       "focus": "Cybersecurity, data management",   "risk": "Med-High",  "notes": "Under active long-term evaluation. Rubrik is a cybersecurity and data management company focused on data resilience — protecting enterprise data from ransomware, ensuring recoverability, and managing data across hybrid cloud environments. As AI adoption accelerates, the volume and value of enterprise data explodes, making data protection and recovery increasingly critical. Rubrik has a subscription-based revenue model (good for predictability) and is growing fast. The risk: the cybersecurity market is crowded and competitive, and the valuation already prices in significant growth. Watch for: revenue acceleration, margin improvement, and any signs of customer concentration. Add on weakness."},
            {"ticker": "NVO",   "company": "Novo Nordisk",     "focus": "GLP-1 drugs, diabetes & obesity",  "risk": "Medium",    "notes": "Under active long-term evaluation. Novo Nordisk is the global leader in GLP-1 drugs (Ozempic, Wegovy) for diabetes and obesity. The obesity drug market is potentially the largest pharmaceutical market ever — hundreds of millions of people globally are candidates for GLP-1 treatment, and early data suggests these drugs also reduce cardiovascular events, sleep apnea, and other conditions beyond weight loss. Novo has the manufacturing scale, the brand recognition, and the clinical data to dominate this market for years. The stock has pulled back significantly from its highs as competition from Eli Lilly (Mounjaro/Zepbound) has intensified — which creates a potentially attractive entry point. A long-term compounder in healthcare, completely uncorrelated to the AI trade."},
        ],
    },
    {
        "name": "📝 Research Notes",
        "tab_color": "1B2A4A",
        "color": "1B2A4A",
        "description": "",
        "stocks": [],
    },
]

# ── Column definitions for tracker sheets ────────────────────────────────────
COLS = [
    ("Ticker",              10,  False),   # 1  A
    ("Company",             28,  False),   # 2  B
    ("Focus / Niche",       30,  False),   # 3  C
    ("Risk",                11,  False),   # 4  D
    ("Current\nPrice",      12,  True),    # 5  E - yellow input
    ("Today's\n+/- %",      11,  False),   # 6  F - green/red auto
    ("Purchase\nPrice",     12,  True),    # 7  G - yellow input
    ("Shares\nOwned",       10,  True),    # 8  H - yellow input
    ("Total\nCost",         13,  False),   # 9  I = G*H
    ("Current\nValue",      13,  False),   # 10 J = E*H
    ("P&L ($)",             13,  False),   # 11 K = J-I
    ("P&L (%)",             10,  False),   # 12 L = K/I
    ("52W\nHigh",           11,  True),    # 13 M - yellow input
    ("52W\nLow",            11,  True),    # 14 N - yellow input
    ("% from\n52W High",    12,  False),   # 15 O = E/M-1
    ("Analyst\nTarget",     12,  True),    # 16 P - yellow input
    ("Upside to\nTarget",   12,  False),   # 17 Q = P/E-1
]

def col_letter(idx):  # 1-based
    return get_column_letter(idx)

def apply_header_row(ws, row, cat_color, names_widths):
    for ci, (name, width, is_input) in enumerate(names_widths, 1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.fill = fill(cat_color)
        cell.font = font(bold=True, color=WHITE, size=10)
        cell.alignment = center()
        cell.border = hdr_border()

def write_stock_row(ws, row, stock, cat_color):
    tickers_col  = 1   # A
    company_col  = 2   # B
    focus_col    = 3   # C
    risk_col     = 4   # D
    cur_col      = 5   # E - current price (yellow input)
    change_col   = 6   # F - today's +/-% (green/red auto)
    buy_col      = 7   # G - purchase price (yellow input)
    shares_col   = 8   # H - shares (yellow input)
    cost_col     = 9   # I = G*H
    val_col      = 10  # J = E*H
    pnl_col      = 11  # K = J-I
    pnlp_col     = 12  # L = K/I
    hi_col       = 13  # M - 52W High (yellow input)
    lo_col       = 14  # N - 52W Low (yellow input)
    fromhi_col   = 15  # O = E/M-1
    tgt_col      = 16  # P - analyst target (yellow input)
    upside_col   = 17  # Q = P/E-1

    bg = LIGHT_GRAY if row % 2 == 0 else WHITE

    def cell(col, value=None, formula=None, fmt=None, is_input=False):
        c = ws.cell(row=row, column=col, value=value if formula is None else formula)
        c.border = THIN_BORDER
        c.alignment = left() if col in (company_col, focus_col) else center()
        if is_input:
            c.fill = fill(YELLOW_IN)
        else:
            c.fill = fill(bg)
        if fmt:
            c.number_format = fmt
        return c

    # Static data
    c = cell(tickers_col, stock["ticker"])
    c.font = Font(bold=True, color="000070C0", name="Calibri", size=11)
    cell(company_col, stock["company"])
    cell(focus_col,   stock["focus"])
    cell(risk_col,    stock["risk"])

    # Current price — pre-filled with live data (yellow so user can override)
    live_price = stock.get("live_price", 0.00)
    cell(cur_col, live_price, is_input=True, fmt='"$"#,##0.00')

    # Today's +/-% — store as e.g. 5.27 (not 0.0527), display with literal %
    # so Excel does NOT apply its built-in ×100 multiplication
    daily_chg = stock.get("daily_change", None)
    chg_cell = ws.cell(row=row, column=change_col)
    chg_cell.border = THIN_BORDER
    chg_cell.alignment = center()
    if daily_chg is not None:
        pct_value = daily_chg * 100   # e.g. 0.052728 → 5.2728
        chg_cell.value = pct_value
        # '+0.00"%"' uses a literal % so no auto-multiplication
        chg_cell.number_format = '+0.00"%";-0.00"%";0.00"%"'
        if pct_value > 0:
            chg_cell.fill = fill("375623")   # dark green bg
            chg_cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        elif pct_value < 0:
            chg_cell.fill = fill("9C0006")   # dark red bg
            chg_cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        else:
            chg_cell.fill = fill(bg)
            chg_cell.font = Font(bold=True, color="595959", size=11, name="Calibri")
    else:
        chg_cell.value = None
        chg_cell.number_format = '0.00"%"'
        chg_cell.fill = fill(bg)

    cell(buy_col,    0.00, is_input=True, fmt='"$"#,##0.00')
    cell(shares_col, 0,    is_input=True, fmt='#,##0')

    r = row
    # Formulas — note column letters shifted by 1
    cell(cost_col,   formula=f"=G{r}*H{r}",             fmt='"$"#,##0.00')
    cell(val_col,    formula=f"=E{r}*H{r}",              fmt='"$"#,##0.00')
    cell(pnl_col,    formula=f"=J{r}-I{r}",              fmt='"$"#,##0.00')
    cell(pnlp_col,   formula=f'=IFERROR(K{r}/I{r},"")',  fmt='0.00%')

    cell(hi_col,     0.00, is_input=True, fmt='"$"#,##0.00')
    cell(lo_col,     0.00, is_input=True, fmt='"$"#,##0.00')
    cell(fromhi_col, formula=f'=IFERROR(E{r}/M{r}-1,"")', fmt='0.00%')

    cell(tgt_col,    0.00, is_input=True, fmt='"$"#,##0.00')
    cell(upside_col, formula=f'=IFERROR(P{r}/E{r}-1,"")',  fmt='0.00%')


def build_tracker_sheet(wb, cat, fetched_at=""):
    ws = wb[cat["name"]]
    ws.sheet_properties.tabColor = cat["tab_color"]
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 32

    cat_color = cat["color"]

    # Row 1 — Category title
    ws.merge_cells(f"A1:{col_letter(len(COLS))}1")
    title_cell = ws["A1"]
    title_cell.value = cat["name"]
    title_cell.fill = fill(cat_color)
    title_cell.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    title_cell.alignment = center()

    # Row 2 — Description
    ws.merge_cells(f"A2:{col_letter(len(COLS))}2")
    desc_cell = ws["A2"]
    desc_cell.value = cat["description"]
    desc_cell.fill = fill("D9E1F2")
    desc_cell.font = Font(italic=True, color=DARK_BLUE, size=10, name="Calibri")
    desc_cell.alignment = left()

    # Row 3 — Legend
    ws.merge_cells(f"A3:{col_letter(len(COLS))}3")
    leg = ws["A3"]
    leg.value = f"🟡 Yellow cells = your inputs  |  All other cells auto-calculate  |  📡 Prices fetched: {fetched_at}"
    leg.fill = fill(YELLOW_IN)
    leg.font = Font(italic=True, color="7F6000", size=9, name="Calibri")
    leg.alignment = center()

    # Row 4 — Column headers
    apply_header_row(ws, 4, cat_color, COLS)

    # Stock rows
    first_data = 5
    for i, stock in enumerate(cat["stocks"]):
        write_stock_row(ws, first_data + i, stock, cat_color)

    last_data = first_data + len(cat["stocks"]) - 1

    # Summary totals row
    if cat["stocks"]:
        tot_row = last_data + 2
        ws.row_dimensions[tot_row].height = 20
        ws.merge_cells(f"A{tot_row}:G{tot_row}")
        tot_label = ws[f"A{tot_row}"]
        tot_label.value = "PORTFOLIO TOTALS"
        tot_label.fill = fill(cat_color)
        tot_label.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        tot_label.alignment = center()

        for ci, col_name in [(8, "Total Cost"), (9, "Current Value"), (10, "P&L ($)")]:
            c = ws.cell(row=tot_row, column=ci)
            cl = col_letter(ci)
            c.value = f"=SUM({cl}{first_data}:{cl}{last_data})"
            c.fill = fill(cat_color)
            c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
            c.alignment = center()
            c.number_format = '"$"#,##0.00'
            c.border = hdr_border()

        pnlp_tot = ws.cell(row=tot_row, column=11)
        pnlp_tot.value = f"=IFERROR(J{tot_row}/H{tot_row},\"\")"
        pnlp_tot.fill = fill(cat_color)
        pnlp_tot.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        pnlp_tot.alignment = center()
        pnlp_tot.number_format = "0.00%"
        pnlp_tot.border = hdr_border()

        # Conditional formatting P&L $ column
        green_rule = CellIsRule(operator="greaterThan", formula=["0"],
                                fill=fill(GREEN_BG), font=Font(color=GREEN_FONT, bold=True, name="Calibri"))
        red_rule   = CellIsRule(operator="lessThan",   formula=["0"],
                                fill=fill(RED_BG),   font=Font(color=RED_FONT,   bold=True, name="Calibri"))
        rng_pnl  = f"J{first_data}:J{last_data}"
        rng_pnlp = f"K{first_data}:K{last_data}"
        ws.conditional_formatting.add(rng_pnl,  green_rule)
        ws.conditional_formatting.add(rng_pnl,  red_rule)
        ws.conditional_formatting.add(rng_pnlp, green_rule)
        ws.conditional_formatting.add(rng_pnlp, red_rule)

        # ── Notes section below totals ─────────────────────────────────────
        # Strategy: NO merged cells, NO wrap_text dependency.
        # Column A (width 14) = ticker label.
        # Column B (overridden to width 130 below) = note text, one wrapped
        # line per row. Plain text cells — works in every viewer.
        notes_start = tot_row + 2

        # Section header in column A
        nh = ws.cell(row=notes_start, column=1)
        nh.value = "INVESTMENT THESIS & RESEARCH NOTES"
        nh.fill = fill(cat_color)
        nh.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        nh.alignment = Alignment(horizontal="left", vertical="center")
        # Extend the header colour across all columns manually
        for ci in range(2, len(COLS) + 1):
            hc = ws.cell(row=notes_start, column=ci)
            hc.fill = fill(cat_color)
        ws.row_dimensions[notes_start].height = 22

        cur_row = notes_start + 1

        NOTE_WIDTH = 130   # chars per line; matches column B override below

        for j, stock in enumerate(cat["stocks"]):
            bg_note = LIGHT_GRAY if j % 2 == 0 else WHITE
            note_text = stock.get("notes", "")

            # ── Label row: col A = ticker, col B = company + risk ─────────
            label_row = cur_row
            la = ws.cell(row=label_row, column=1, value=stock["ticker"])
            la.fill = fill("D9E1F2")
            la.font = Font(bold=True, color=DARK_BLUE, size=11, name="Calibri")
            la.alignment = Alignment(horizontal="center", vertical="center")

            lb = ws.cell(row=label_row, column=2,
                         value=f"  {stock['company']}   |   Risk: {stock['risk']}")
            lb.fill = fill("D9E1F2")
            lb.font = Font(bold=True, color=DARK_BLUE, size=11, name="Calibri")
            lb.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[label_row].height = 20
            cur_row += 1

            # ── Note lines: one plain-text row per wrapped line ───────────
            lines = textwrap.wrap(note_text, width=NOTE_WIDTH) if note_text else ["(no notes)"]
            for line in lines:
                nr = cur_row
                # Col A: empty coloured cell
                ws.cell(row=nr, column=1).fill = fill(bg_note)
                # Col B: the line of text
                nc = ws.cell(row=nr, column=2, value="  " + line)
                nc.fill = fill(bg_note)
                nc.font = Font(color="1B2A4A", size=10, name="Calibri")
                nc.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[nr].height = 16
                cur_row += 1

            # Small visual gap between stocks
            ws.cell(row=cur_row, column=1).fill = fill(WHITE)
            ws.cell(row=cur_row, column=2).fill = fill(WHITE)
            ws.row_dimensions[cur_row].height = 6
            cur_row += 1

    # Column widths
    for ci, (_, width, _) in enumerate(COLS, 1):
        ws.column_dimensions[col_letter(ci)].width = width
    # Column B overridden wide so notes text is fully visible (no merging needed)
    ws.column_dimensions["B"].width = 130

    return ws


def build_dashboard(wb, categories, fetched_at=""):
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_properties.tabColor = "1B2A4A"
    ws.freeze_panes = "A6"

    # Title banner
    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value = "📈 PERSONAL STOCK TRACKER — PORTFOLIO DASHBOARD"
    t.fill = fill(DARK_BLUE)
    t.font = Font(bold=True, color="FFD700", size=16, name="Calibri")
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:P2")
    sub = ws["A2"]
    sub.value = "Update prices in each category tab. This dashboard auto-totals across all categories."
    sub.fill = fill(MID_BLUE)
    sub.font = Font(italic=True, color=WHITE, size=10, name="Calibri")
    sub.alignment = center()
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:P3")
    leg3 = ws["A3"]
    leg3.value = f"🟡 Yellow cells = your inputs  |  All other cells auto-calculate  |  Green = gain  |  Red = loss  |  📡 Prices fetched: {fetched_at}"
    leg3.fill = fill(YELLOW_IN)
    leg3.font = Font(italic=True, color="7F6000", size=9, name="Calibri")
    leg3.alignment = center()
    ws.row_dimensions[3].height = 16

    ws.row_dimensions[4].height = 8

    # Summary header row 5
    dash_cols = ["Category", "# Stocks", "Risk Profile", "Theme", "Total Cost", "Current Value", "P&L ($)", "P&L (%)"]
    dash_widths = [26, 10, 14, 38, 16, 16, 14, 10]
    for ci, (col_name, w) in enumerate(zip(dash_cols, dash_widths), 1):
        c = ws.cell(row=5, column=ci, value=col_name)
        c.fill = fill(DARK_BLUE)
        c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        c.alignment = center()
        c.border = hdr_border()
        ws.column_dimensions[col_letter(ci)].width = w

    ws.row_dimensions[5].height = 22

    themes = {
        "⚛️ Quantum Computing":      "High-risk quantum plays — 10× potential",
        "🖥️ Semiconductors & Compute": "Core compute names — stability & compounders",
        "⚙️ AI Picks & Shovels":       "Smaller-cap NVDA-adjacent — asymmetric upside",
        "💡 Photonics & Optics":       "Optical interconnect — key AI DC bottleneck",
        "🏗️ DC Builders & Cooling":    "Data center construction & cooling plays",
        "⚡ Power & Energy":            "Power generation, grid & nuclear for AI buildout",
        "☁️ Neoclouds & AI Software":  "Next-gen cloud & government AI software",
        "🛡️ Defense & Drones":         "$1.5T defense budget — drone & defense tech",
        "🚀 Space Economy":            "Space infrastructure — SpaceX IPO catalyst",
        "🔭 Long Term Watchlist":      "Research candidates under active evaluation",
    }
    risk_profiles = {
        "⚛️ Quantum Computing":      "Very High",
        "🖥️ Semiconductors & Compute": "Medium",
        "⚙️ AI Picks & Shovels":       "High",
        "💡 Photonics & Optics":       "High",
        "🏗️ DC Builders & Cooling":    "Low-Medium",
        "⚡ Power & Energy":            "Medium-High",
        "☁️ Neoclouds & AI Software":  "Medium-High",
        "🛡️ Defense & Drones":         "Very High",
        "🚀 Space Economy":            "High",
        "🔭 Long Term Watchlist":      "Mixed",
    }

    data_cats = [c for c in categories if c["stocks"]]
    summary_rows = []

    for i, cat in enumerate(data_cats):
        row = 6 + i
        ws.row_dimensions[row].height = 20
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        cat_name = cat["name"]
        n_stocks = len(cat["stocks"])
        theme = themes.get(cat_name, "")
        risk = risk_profiles.get(cat_name, "")

        cells_data = [cat_name, n_stocks, risk, theme, "", "", "", ""]
        for ci, val in enumerate(cells_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill(bg)
            c.border = THIN_BORDER
            c.alignment = left() if ci in (1, 4) else center()
            if ci == 1:
                c.font = Font(bold=True, color="00" + cat["color"], name="Calibri", size=11)

        summary_rows.append(row)

    # Grand total row
    tot_row = 6 + len(data_cats) + 1
    ws.row_dimensions[tot_row].height = 24
    ws.merge_cells(f"A{tot_row}:D{tot_row}")
    gt = ws[f"A{tot_row}"]
    gt.value = "GRAND TOTAL — ALL CATEGORIES"
    gt.fill = fill(DARK_BLUE)
    gt.font = Font(bold=True, color="FFD700", size=12, name="Calibri")
    gt.alignment = center()
    for ci in range(1, 9):
        c = ws.cell(row=tot_row, column=ci)
        c.fill = fill(DARK_BLUE)
        c.border = hdr_border()

    # Key notes section
    note_row = tot_row + 3
    ws.merge_cells(f"A{note_row}:H{note_row}")
    nh = ws[f"A{note_row}"]
    nh.value = "📌  KEY RESEARCH NOTES & THESIS"
    nh.fill = fill(MID_BLUE)
    nh.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
    nh.alignment = left()
    ws.row_dimensions[note_row].height = 22

    notes_text = [
        ("To-Do Actions",
         "• Transfer 401k from Accrue/Vestwell → Charles Schwab\n• Transfer 401k from Vanguard → Charles Schwab\n• Open Capital One Savings Account (for holding taxes)\n• Learn Unusual Whales website (watch YouTube tutorials)"),
        ("AI Infrastructure Thesis",
         "The AI trade is no longer just about GPUs. It is shifting down stack into:\n  Memory Names | Photonics & Optics | CPU Makers (INTC, AMD, ARM, NVDA)\n  Data Center Builders & Materials (CAT, FIX, POWL) | Cooling (VRT, FIX)\n  Power Generation & Grid (GEV, BE, PWR) | Nuclear (OKLO)"),
        ("What Connects Small-Cap AI Names",
         "These are not mini-NVDA clones. They represent:\n  1. AI Infrastructure Bottlenecks: power (NVTS), chip testing (AEHR), data movement (CRDO/MTSI), efficiency (ALGM)\n  2. Early Monetization Phase: revenue exists but still scaling\n  3. Narrative Sensitivity: move on earnings beats, guidance raises, AI capex headlines"),
        ("Recommended Core Portfolio Mix",
         "LRCX (stability) | DRAM ETF | LITE (photonics moonshot) | NVDA (stability) | AVGO (stability)\nFIX or POWL or CAT (pick one) | GEV (stability) | BE (if hydrogen thesis) | OKLO (nuclear moonshot, small %)\nNBIS (future hyperscaler) | PLTR (gov AI software) | ONDS (drone defense, speculative) | RKLB (space economy)"),
        ("Quantum Run Note",
         "Next run will be quantum. Buy IONQ, RGTI, and QBTS. Easy 10× within 2 years."),
        ("Three Outliers Outside Data Center Trade",
         "PLTR — AI Software. Integral to entire government AI stack.\nONDS — Drones & Defense. $1.5T defense budget, $56B specifically to drones via DAWG Program.\nRKLB — Space Economy. Huge future TAM. SpaceX IPO in June is a near-term catalyst."),
    ]

    for j, (heading, body) in enumerate(notes_text):
        h_row = note_row + 1 + j * 4
        ws.merge_cells(f"A{h_row}:H{h_row}")
        hc = ws[f"A{h_row}"]
        hc.value = heading
        hc.fill = fill("D9E1F2")
        hc.font = Font(bold=True, color=DARK_BLUE, size=10, name="Calibri")
        hc.alignment = left()
        ws.row_dimensions[h_row].height = 18

        b_row = h_row + 1
        ws.merge_cells(f"A{b_row}:H{b_row + 2}")
        bc = ws[f"A{b_row}"]
        bc.value = body
        bc.fill = fill(WHITE)
        bc.font = Font(color="1B2A4A", size=9, name="Calibri")
        bc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        bc.border = THIN_BORDER
        ws.row_dimensions[b_row].height = 60

    return ws


def build_notes_sheet(wb):
    ws = wb["📝 Research Notes"]
    ws.sheet_properties.tabColor = "1B2A4A"

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "📝 RESEARCH NOTES & INVESTMENT THESIS"
    t.fill = fill(DARK_BLUE)
    t.font = Font(bold=True, color="FFD700", size=14, name="Calibri")
    t.alignment = center()
    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    notes = [
        ("401k To-Do",            "• Transfer 401k from Accrue (Now Vestwell) → Charles Schwab\n• Transfer 401k from Vanguard (1-800-523-1188) → Charles Schwab"),
        ("Savings Account",       "• Open Savings Account with Capital One (for holding taxes)"),
        ("Research Tools",        "• Learn how to use 'Unusual Whales' website — watch YouTube tutorials"),
        ("Quantum Thesis",        "Next run will be of quantum. Buy IONQ, RGTI, and QBTS. Easy 10× within 2 years."),
        ("NVTS",                  "Navitas Semiconductor — GaN power chips. AI data centers need radically more efficient power delivery. NVDA parallel: picks-and-shovels for compute growth. High-risk bet on next-gen power architecture shift."),
        ("AEHR",                  "AEHR Test Systems — Semiconductor burn-in and test equipment. AI chips require more advanced testing at scale. More chips → more testing bottleneck demand. Classic shovels in a gold rush setup."),
        ("CRDO",                  "Credo Technology — High-speed connectivity / SerDes for AI data centers. GPUs don't matter if data can't move between them. Direct exposure to AI cluster networking buildout. One of the cleanest AI infrastructure picks-and-shovels names."),
        ("ALGM",                  "Allegro MicroSystems — Power & sensing semiconductors. Electrification + AI hardware power efficiency. Steady but still early in AI rerating narrative. More of a quiet compounder than hype stock."),
        ("HIMX",                  "Himax Technologies — Display drivers + AR/AI vision chips. Exposure to AI edge devices + optics. Often overlooked ADR with cyclical breakout behavior. Higher volatility, but historically rerates fast on cycles."),
        ("MTSI",                  "MACOM — RF, microwave, high-speed analog chips. AI datacenter interconnect + telecom backbone. Beneficiary of long-term bandwidth expansion trend. Less hype, more industrial AI backbone exposure."),
        ("AEVA",                  "Aeva Technologies — Lidar + sensing tech. Autonomous systems + robotics long tail. Still early commercialization phase. Pure asymmetric bet: low probability, high upside."),
        ("AI Infrastructure",     "The AI trade is shifting down stack: Memory Names | Photonics & Optics | CPU Makers (INTC, AMD, ARM, NVDA) | Actual Builders & Materials (CAT, FIX, POWL) | Cooling Companies (VRT, FIX) | Power Generation & Grid (GEV, BE, PWR)"),
        ("LITE",                  "Lumentum — King of Photonics. Could moonshot. Critical infrastructure for AI data movement."),
        ("OKLO",                  "Nuclear SMRs and Nuclear Fuel Waste. Huge moonshot potential but big regulatory and execution risk. Worth a small percentage of portfolio because of its massive potential."),
        ("NBIS",                  "Nebius Group — Leading Neocloud that also owns Clickhouse and AV Ride. This is a modern day Amazon or Google style company — main product plus several other revenue streams that could be massive in the future. I think this will be a future Hyperscaler like GOOGL, Amazon, META."),
        ("PLTR",                  "Palantir — AI Software. So integral to the entire government AI stack and is not going anywhere soon. Outlier outside of the data center trade."),
        ("ONDS",                  "Ondas Holdings — Drones and defense tech. Government just announced a new $1.5 Trillion defense budget with $56 Billion going specifically to drones and counter-drones through the DAWG Program. Very speculative but best chance to scale after their recent acquisitions and partnerships made in March 2026."),
        ("RKLB",                  "Rocket Lab — Space Economy and Space Infrastructure. Huge future potential TAM over the next decade or two. Speculative for sure but huge upside potential that will compound over time if they can execute. Brilliant CEO. Will run into the SpaceX IPO in June in the near term."),
        ("Recommended Portfolio", "LRCX (stability) | DRAM ETF | LITE (photonics moonshot) | NVDA (stability) | AVGO (stability) | FIX or POWL or CAT (pick one) | GEV (stability) | BE (hydrogen moonshot) | OKLO (nuclear moonshot, small %) | NBIS (future hyperscaler)"),
    ]

    for i, (heading, body) in enumerate(notes, 2):
        ws.row_dimensions[i].height = max(15, body.count("\n") * 15 + 20)
        hc = ws.cell(row=i, column=1, value=heading)
        hc.fill = fill("D9E1F2" if i % 2 == 0 else LIGHT_GRAY)
        hc.font = Font(bold=True, color=DARK_BLUE, size=10, name="Calibri")
        hc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        hc.border = THIN_BORDER

        bc = ws.cell(row=i, column=2, value=body)
        bc.fill = fill(WHITE)
        bc.font = Font(color="1B2A4A", size=9, name="Calibri")
        bc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        bc.border = THIN_BORDER


# ── Fetch live prices ─────────────────────────────────────────────────────────
def fetch_live_prices(categories):
    all_tickers = []
    for cat in categories:
        for stock in cat.get("stocks", []):
            all_tickers.append(stock["ticker"])

    print(f"Fetching live prices + today's change for {len(all_tickers)} tickers...")
    prices = {}
    daily_changes = {}
    try:
        tickers_obj = yf.Tickers(" ".join(all_tickers))
        for t in all_tickers:
            try:
                fi         = tickers_obj.tickers[t].fast_info
                last_price = fi.last_price
                prev_close = fi.previous_close
                if last_price:
                    prices[t] = round(float(last_price), 2)
                if last_price and prev_close and prev_close != 0:
                    daily_changes[t] = float((last_price - prev_close) / prev_close)
            except Exception:
                pass
        print(f"  ✓ Got prices for {len(prices)} tickers, daily changes for {len(daily_changes)}")
    except Exception as e:
        print(f"  ⚠ Price fetch failed: {e}. Using 0.00 placeholders.")

    fetched_at = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    return prices, daily_changes, fetched_at


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    # Fetch live prices + daily changes
    live_prices, daily_changes, fetched_at = fetch_live_prices(CATEGORIES)

    # Inject prices and daily changes into stock dicts
    for cat in CATEGORIES:
        for stock in cat.get("stocks", []):
            t = stock["ticker"]
            stock["live_price"]   = round(live_prices.get(t, 0.00), 2)
            stock["daily_change"] = daily_changes.get(t, None)

    wb = openpyxl.Workbook()

    # Pre-create all sheets so ordering is correct
    for cat in CATEGORIES:
        if cat["name"] == "📊 Dashboard":
            continue
        wb.create_sheet(cat["name"])

    build_dashboard(wb, CATEGORIES, fetched_at)

    for cat in CATEGORIES:
        if cat["name"] in ("📊 Dashboard", "📝 Research Notes"):
            continue
        if cat["stocks"]:
            build_tracker_sheet(wb, cat, fetched_at)

    build_notes_sheet(wb)

    # Move dashboard to front
    wb.move_sheet("📊 Dashboard", offset=-len(wb.sheetnames))

    out_path = "/home/runner/workspace/Financial_Stock_Tracker.xlsx"
    wb.save(out_path)
    print(f"Saved → {out_path}")

main()
